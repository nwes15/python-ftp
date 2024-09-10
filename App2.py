import flet as ft
from openpyxl import Workbook, load_workbook
from ftplib import FTP
import os

# Função para criar a planilha
def criar_planilha(nome_arquivo):
    arquivox = Workbook()
    plan0 = arquivox.active
    plan0.title = "ImportTasks"

    plan0['A1'] = 'Nº Tarefa'
    plan0['B1'] = 'Descrição'
    plan0['C1'] = 'Tipo tarefa'
    plan0['D1'] = 'Nome cliente'
    plan0['E1'] = 'Rua'
    plan0['F1'] = "Número do local"
    plan0['G1'] = "Bairro"
    plan0['H1'] = "Telefone"
    plan0['I1'] = 'Longitude'
    plan0['J1'] = 'Latitude'

    arquivox.save(f"{nome_arquivo}.xlsx")

# Função para adicionar uma linha
def adicionar_linha(nome_arquivo, numero_tarefa, descricao, tipo_tarefa, nome_cliente, rua, numero_local, bairro, telefone, longitude, latitude):
    # Verificar se o arquivo já existe; se não, criar a planilha
    if not os.path.exists(f"{nome_arquivo}.xlsx"):
        criar_planilha(nome_arquivo)
        
    arquivox = load_workbook(f"{nome_arquivo}.xlsx")
    plan0 = arquivox.active

    # Adicionar os dados na próxima linha
    plan0.append([numero_tarefa, descricao, tipo_tarefa, nome_cliente, rua, numero_local, bairro, telefone, longitude, latitude])
    arquivox.save(f"{nome_arquivo}.xlsx")

# Função para enviar o arquivo via FTP
def enviar_ftp(nome_arquivo, host, user, senha):
    try:
        with FTP(host) as ftp:
            ftp.login(user=user, passwd=senha)
            with open(f"{nome_arquivo}.xlsx", "rb") as file:
                ftp.storbinary(f"STOR {nome_arquivo}.xlsx", file)
        return "Arquivo enviado com sucesso!"
    except Exception as e:
        return f"Erro ao enviar o arquivo: {e}"

# Função principal do app
def main(page: ft.Page):

    # Variável para rastrear se o arquivo já foi criado
    arquivo_criado = False

    # Campos de input para os dados
    input_nome_arquivo = ft.TextField(label="Nome do arquivo (sem extensão)")
    input_numero_tarefa = ft.TextField(label="Nº Tarefa")
    input_descricao = ft.TextField(label="Descrição")
    input_tipo_tarefa = ft.TextField(label="Tipo de tarefa")
    input_nome_cliente = ft.TextField(label="Nome do cliente")
    input_rua = ft.TextField(label="Rua")
    input_numero_local = ft.TextField(label="Número do local")
    input_bairro = ft.TextField(label="Bairro")
    input_telefone = ft.TextField(label="Telefone")
    input_longitude = ft.TextField(label="Longitude")
    input_latitude = ft.TextField(label="Latitude")

    # Função para adicionar linha
    def btn_add_linha(e):
        nonlocal arquivo_criado
        
        # Se o arquivo ainda não foi criado, cria o arquivo
        if not arquivo_criado:
            criar_planilha(input_nome_arquivo.value)
            arquivo_criado = True
            # Desabilitar o campo do nome do arquivo após a criação
            input_nome_arquivo.disabled = True
            page.update()

        # Adicionar uma nova linha à planilha
        adicionar_linha(
            input_nome_arquivo.value,
            input_numero_tarefa.value,
            input_descricao.value,
            input_tipo_tarefa.value,
            input_nome_cliente.value,
            input_rua.value,
            input_numero_local.value,
            input_bairro.value,
            input_telefone.value,
            input_longitude.value,
            input_latitude.value,
        )

        # Limpar os campos após adicionar a linha
        input_numero_tarefa.value = ""
        input_descricao.value = ""
        input_tipo_tarefa.value = ""
        input_nome_cliente.value = ""
        input_rua.value = ""
        input_numero_local.value = ""
        input_bairro.value = ""
        input_telefone.value = ""
        input_longitude.value = ""
        input_latitude.value = ""
        
        page.snack_bar = ft.SnackBar(ft.Text("Linha adicionada e campos limpos com sucesso!"))
        page.snack_bar.open = True
        page.update()

    # Função para salvar a planilha e ir para a tela de FTP
    def btn_salvar_planilha(e):
        if not os.path.exists(f"{input_nome_arquivo.value}.xlsx"):
            criar_planilha(input_nome_arquivo.value)
        page.snack_bar = ft.SnackBar(ft.Text("Planilha criada com sucesso! Indo para a tela de FTP..."))
        page.snack_bar.open = True
        page.update()
        tela_ftp(page)  # Ir para a tela de FTP após salvar a planilha

    # Função para abrir a tela de FTP
    def tela_ftp(page: ft.Page):
        # Campos de input para FTP
        input_ftp_host = ft.TextField(label="FTP Host")
        input_ftp_user = ft.TextField(label="FTP Usuário")
        input_ftp_senha = ft.TextField(label="FTP Senha")

        # Função para enviar o arquivo via FTP
        def btn_enviar_ftp(e):
            resultado = enviar_ftp(input_nome_arquivo.value, input_ftp_host.value, input_ftp_user.value, input_ftp_senha.value)
            page.snack_bar = ft.SnackBar(ft.Text(resultado))
            page.snack_bar.open = True
            page.update()

        # Função para voltar para a tela principal
        def btn_voltar(e):
            page.controls.clear()
            main(page)

        # Adicionando botões para enviar FTP e voltar
        btn_ftp = ft.ElevatedButton(text="Enviar por FTP", on_click=btn_enviar_ftp)
        btn_voltar = ft.ElevatedButton(text="Voltar", on_click=btn_voltar)

        # Limpa a tela e adiciona os novos elementos
        page.controls.clear()
        page.add(
            input_ftp_host,
            input_ftp_user,
            input_ftp_senha,
            ft.Row([btn_ftp, btn_voltar]),
        )
        page.update()

    # Botões
    btn_add = ft.ElevatedButton(text="Adicionar Linha", on_click=btn_add_linha)
    btn_save = ft.ElevatedButton(text="Salvar Planilha e Enviar FTP", on_click=btn_salvar_planilha)

    # Adicionar os elementos à página principal
    page.add(
        input_nome_arquivo,
        input_numero_tarefa,
        input_descricao,
        input_tipo_tarefa,
        input_nome_cliente,
        input_rua,
        input_numero_local,
        input_bairro,
        input_telefone,
        input_longitude,
        input_latitude,
        ft.Row([btn_add, btn_save]),
    )

# Executar o app
ft.app(target=main)